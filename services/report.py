import io
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from db.database import get_db

_PERIOD_LABELS = {
    "today": "Today",
    "daily": "Today",
    "week": "Last 7 days",
    "weekly": "Last 7 days",
    "month": "This month",
    "monthly": "This month",
}


def _date_range(period: str) -> tuple[date, date, str]:
    today = date.today()
    if period in ("today", "daily"):
        return today, today, today.strftime("%Y-%m-%d")
    if period in ("week", "weekly"):
        start = today - timedelta(days=6)
        return start, today, f"{start.strftime('%Y-%m-%d')}_to_{today.strftime('%Y-%m-%d')}"
    # month / monthly (default)
    start = date(today.year, today.month, 1)
    return start, today, today.strftime("%Y-%m")


def _header_row(ws, headers: list[str]) -> None:
    ws.append(headers)
    fill = PatternFill("solid", fgColor="2E75B6")
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)


async def build_report(
    user_id: int, vehicle_id: int, period: str, currency: str,
    cleared_at=None,
) -> tuple[bytes, str]:
    start, end, label = _date_range(period)

    effective_start = start
    if cleared_at:
        clear_date = cleared_at.date() if hasattr(cleared_at, "date") else cleared_at
        if clear_date > start:
            effective_start = clear_date

    async with get_db() as db:
        trips = await db.fetch(
            """SELECT t.occurred_at, t.amount, t.destination,
                      p.display_name AS passenger, t.paid, t.payment_method
               FROM trips t
               LEFT JOIN passengers p ON p.id = t.passenger_id
               WHERE t.user_id = $1
                 AND t.occurred_at::date >= $2
                 AND t.occurred_at::date <= $3
                 AND ($4::timestamptz IS NULL OR t.occurred_at >= $4)
               ORDER BY t.occurred_at""",
            user_id, effective_start, end, cleared_at,
        )
        expenses = await db.fetch(
            """SELECT occurred_at, type, amount, note, litres
               FROM expenses
               WHERE user_id = $1
                 AND occurred_at::date >= $2
                 AND occurred_at::date <= $3
                 AND ($4::timestamptz IS NULL OR occurred_at >= $4)
               ORDER BY occurred_at""",
            user_id, effective_start, end, cleared_at,
        )
        remittance = await db.fetch(
            """SELECT paid_on, amount, status
               FROM remittance_log
               WHERE vehicle_id = $1 AND paid_on >= $2 AND paid_on <= $3
                 AND ($4::timestamptz IS NULL OR created_at >= $4)
               ORDER BY paid_on""",
            vehicle_id, effective_start, end, cleared_at,
        )

    wb = openpyxl.Workbook()

    # ── Sheet 1: Trips ───────────────────────────────────────────────────────
    ws_trips = wb.active
    ws_trips.title = "Trips"
    _header_row(ws_trips, ["Date", "Time", f"Amount ({currency})", "Destination", "Client", "Paid", "Method"])
    for r in trips:
        ws_trips.append([
            r["occurred_at"].strftime("%Y-%m-%d"),
            r["occurred_at"].strftime("%H:%M"),
            float(r["amount"]),
            r["destination"] or "",
            r["passenger"] or "",
            "Yes" if r["paid"] else "No",
            r["payment_method"] or "CASH",
        ])
    _auto_width(ws_trips)

    # ── Sheet 2: Expenses ────────────────────────────────────────────────────
    ws_exp = wb.create_sheet("Expenses")
    _header_row(ws_exp, ["Date", "Time", "Type", f"Amount ({currency})", "Note", "Litres"])
    for r in expenses:
        ws_exp.append([
            r["occurred_at"].strftime("%Y-%m-%d"),
            r["occurred_at"].strftime("%H:%M"),
            r["type"].title(),
            float(r["amount"]),
            r["note"] or "",
            float(r["litres"]) if r["litres"] else "",
        ])
    _auto_width(ws_exp)

    # ── Sheet 3: Remittance ──────────────────────────────────────────────────
    ws_remit = wb.create_sheet("Remittance")
    _header_row(ws_remit, ["Date", f"Amount ({currency})", "Status"])
    for r in remittance:
        ws_remit.append([
            r["paid_on"].strftime("%Y-%m-%d"),
            float(r["amount"]) if r["status"] != "REST" else 0.0,
            r["status"].title(),
        ])
    _auto_width(ws_remit)

    # ── Sheet 4: Summary ─────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    gross = sum(float(r["amount"]) for r in trips if r["paid"])
    unpaid_total = sum(float(r["amount"]) for r in trips if not r["paid"])
    costs = sum(float(r["amount"]) for r in expenses)
    remit_paid = sum(float(r["amount"]) for r in remittance if r["status"] == "PAID")
    net = gross - costs - remit_paid

    rows = [
        ("Period", f"{start} → {end}"),
        ("", ""),
        (f"Gross earnings ({currency})", gross),
        (f"Total expenses ({currency})", costs),
        (f"Remittance paid ({currency})", remit_paid),
        ("", ""),
        (f"Net profit ({currency})", net),
        ("", ""),
        ("Trips (paid)", sum(1 for r in trips if r["paid"])),
        ("Trips (unpaid)", sum(1 for r in trips if not r["paid"])),
        (f"Unpaid amount ({currency})", unpaid_total),
    ]
    _header_row(ws_sum, ["Metric", "Value"])
    for label_cell, val in rows:
        ws_sum.append([label_cell, val])

    bold = Font(bold=True)
    for row in ws_sum.iter_rows(min_row=2):
        if row[0].value:
            row[0].font = bold

    # Highlight net profit row
    for row in ws_sum.iter_rows(min_row=2):
        if row[0].value and "Net profit" in str(row[0].value):
            fill = PatternFill("solid", fgColor="E2EFDA" if net >= 0 else "FFDDC1")
            for cell in row:
                cell.fill = fill
                cell.font = Font(bold=True)

    _auto_width(ws_sum)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), f"driver_ledger_{label}.xlsx"
